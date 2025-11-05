#!/usr/bin/env python3
from openpyxl import load_workbook

filename = 'MBC Relatorio_anuncios_patrocinados_setembro ate outubro 2025.xlsx'
wb = load_workbook(filename)

print('Sheet names:', wb.sheetnames)
print()

# Get the first sheet
ws = wb[wb.sheetnames[0]]

print(f'First sheet: {wb.sheetnames[0]}')
print()

# Print first 5 rows
print('First 5 rows:')
for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True)):
    print(f'Row {i+1}:', row)
    print()

# Print column headers (assuming row 1)
print('Column headers:')
headers = [cell.value for cell in ws[1]]
for i, header in enumerate(headers):
    if header:
        print(f'  Column {i+1}: {header}')
